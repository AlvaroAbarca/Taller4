import matplotlib.pyplot as plt
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl import load_workbook

ANIOS = [1998,1999,2000,2001,2002,
         2003,2004,2005,2006,2007,
         2008,2009,2010,2011,2012,
         2013,2014,2015]

PROM = {"1998":None,
        "1999":None,
        "2000":None,
        "2001":None,
        "2002":None,
        "2003":None,
        "2004":None,
        "2005":None,
        "2006":None,
        "2007":None,
        "2008":None,
        "2009":None,
        "2010":None,
        "2011":None,
        "2012":None,
        "2013":None,
        "2014":None,
        "2015":None,
        }

################ 1998 ###################
wb = load_workbook(filename = 'DEPTO_4TO_1998.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'N_CDP':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'NU_IRT_M':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'NU_IRT_C':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["1998"] = promedios
wb.close()

################ 1999 ###################
wb = load_workbook(filename = 'DEPTO_4TO_1999.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'DEPTO':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'MATE':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'LENG':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["1999"] = promedios
wb.close()

################ 2000 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2000.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'DEPTO':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'MAT':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'CAST':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2000"] = promedios
wb.close()

################ 2001 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2001.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'DEPTO' or sheet_ranges[indice].value == 'DEPTOS':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'MAT':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'CAST':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2001"] = promedios
wb.close()

################ 2002 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2002.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'DEPTO' or sheet_ranges[indice].value == 'DEPTOS':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'MAT' or sheet_ranges[indice].value == 'MATE':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'CAST'or sheet_ranges[indice].value == 'LENG':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2002"] = promedios
wb.close()

################ 2003 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2003.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'DEPTO' or sheet_ranges[indice].value == 'DEPTOS':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'MAT' or sheet_ranges[indice].value == 'MATE':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'CAST'or sheet_ranges[indice].value == 'LENG':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2003"] = promedios
wb.close()

################ 2004 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2004.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'DEPTO' or sheet_ranges[indice].value == 'DEPTOS':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'MAT' or sheet_ranges[indice].value == 'MATE':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'CAST'or sheet_ranges[indice].value == 'LENG':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2004"] = promedios
wb.close()

################ 2005 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2005.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'DEPTO' or sheet_ranges[indice].value == 'DEPTOS':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'MAT' or sheet_ranges[indice].value == 'MATE' or sheet_ranges[indice].value == 'PROM_MAT':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'CAST'or sheet_ranges[indice].value == 'LENG' or sheet_ranges[indice].value == 'PROM_LEN':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2005"] = promedios
wb.close()

################ 2006 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2006.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'DEPTO' or sheet_ranges[indice].value == 'DEPTOS':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'MAT' or sheet_ranges[indice].value == 'MATE' or sheet_ranges[indice].value == 'PROM_MAT':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'CAST'or sheet_ranges[indice].value == 'LENG' or sheet_ranges[indice].value == 'PROM_LEN':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2006"] = promedios
wb.close()

################ 2007 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2007.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'DEPTO' or sheet_ranges[indice].value == 'DEPTOS':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'MAT' or sheet_ranges[indice].value == 'MATE' or sheet_ranges[indice].value == 'PROM_MAT':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'CAST'or sheet_ranges[indice].value == 'LENG' or sheet_ranges[indice].value == 'PROM_LEN':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2007"] = promedios
wb.close()

################ 2008 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2008.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'DEPTO' or sheet_ranges[indice].value == 'DEPTOS':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'MAT' or sheet_ranges[indice].value == 'MATE' or sheet_ranges[indice].value == 'PROM_MAT':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'CAST'or sheet_ranges[indice].value == 'LENG' or sheet_ranges[indice].value == 'PROM_LEN' or sheet_ranges[indice].value == 'PROM_LECT':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2008"] = promedios
wb.close()

################ 2009 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2009.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'DEPTO' or sheet_ranges[indice].value == 'DEPTOS':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'MAT' or sheet_ranges[indice].value == 'MATE' or sheet_ranges[indice].value == 'PROM_MAT':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'CAST'or sheet_ranges[indice].value == 'LENG' or sheet_ranges[indice].value == 'PROM_LEN' or sheet_ranges[indice].value == 'PROM_LECT':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2009"] = promedios
wb.close()

################ 2010 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2010.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'DEPTO' or sheet_ranges[indice].value == 'DEPTOS':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'MAT' or sheet_ranges[indice].value == 'MATE' or sheet_ranges[indice].value == 'PROM_MAT':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'CAST'or sheet_ranges[indice].value == 'LENG' or sheet_ranges[indice].value == 'PROM_LEN' or sheet_ranges[indice].value == 'PROM_LECT':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2010"] = promedios
wb.close()

################ 2011 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2011.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'DEPTO' or sheet_ranges[indice].value == 'DEPTOS':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'MAT' or sheet_ranges[indice].value == 'MATE' or sheet_ranges[indice].value == 'PROM_MAT':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'CAST'or sheet_ranges[indice].value == 'LENG' or sheet_ranges[indice].value == 'PROM_LEN' or sheet_ranges[indice].value == 'PROM_LECT':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2011"] = promedios
wb.close()

################ 2012 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2012.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'DEPTO' or sheet_ranges[indice].value == 'DEPTOS':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'MAT' or sheet_ranges[indice].value == 'MATE' or sheet_ranges[indice].value == 'PROM_MAT':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'CAST'or sheet_ranges[indice].value == 'LENG' or sheet_ranges[indice].value == 'PROM_LEN' or sheet_ranges[indice].value == 'PROM_LECT':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2012"] = promedios
wb.close()

################ 2013 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2013.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'DEPTO' or sheet_ranges[indice].value == 'DEPTOS' or sheet_ranges[indice].value == 'DEPROV':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'MAT' or sheet_ranges[indice].value == 'MATE' or sheet_ranges[indice].value == 'PROM_MAT':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'CAST'or sheet_ranges[indice].value == 'LENG' or sheet_ranges[indice].value == 'PROM_LEN' or sheet_ranges[indice].value == 'PROM_LECT':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2013"] = promedios
wb.close()

################ 2014 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2014.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'nom_deprov':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'prom_mate4b_deprov':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'prom_lect4b_deprov':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2014"] = promedios
wb.close()

################ 2015 ###################
wb = load_workbook(filename = 'DEPTO_4TO_2015.xlsx')
sheet = wb.sheetnames[0]
sheet_ranges = wb[sheet]
provincia = None 
lenguaje = None
matematica = None

Datos = {"P_L":[],"P_M":[],"O_L":[],"O_M":[]}
promedios = {"P_L":0,"P_M":0,"O_L":0,"O_M":0}

for letra in range(65,91):
    indice = chr(letra)+str(1)
    if sheet_ranges[indice].value == 'nom_deprov':
        provincia = letra-65  #posicion del arreglo donde se almacena la provincia
    if sheet_ranges[indice].value == 'prom_mate4b_deprov':
        matematica = letra-65 #posicion del arreglo donde se almacena el puntaje matematica
    if sheet_ranges[indice].value == 'prom_lect4b_deprov':
        lenguaje = letra-65   #posicion del arreglo donde se almacena el puntaje lenguaje
         

for filas in sheet_ranges.rows:
    if filas[provincia].value == 'Santiago Oriente' or filas[provincia].value == 'SANTIAGO ORIENTE':
        Datos["O_L"].append(filas[lenguaje].value)
        promedios["O_L"] += int(filas[lenguaje].value)
        
        Datos["O_M"].append(filas[matematica].value)
        promedios["O_M"] += int(filas[matematica].value)
        
    if filas[provincia].value == 'Santiago Poniente' or filas[provincia].value =='SANTIAGO PONIENTE':
        Datos["P_L"].append(filas[lenguaje].value)
        promedios["P_L"] += int(filas[lenguaje].value)
        
        Datos["P_M"].append(filas[matematica].value)
        promedios["P_M"] += int(filas[matematica].value)


promedios["O_L"] = promedios["O_L"]/(len(Datos["O_L"]))
promedios["O_M"] = promedios["O_M"]/(len(Datos["O_M"]))
promedios["P_L"] = promedios["P_L"]/(len(Datos["P_L"]))
promedios["P_M"] = promedios["P_M"]/(len(Datos["P_M"]))

PROM["2015"] = promedios
wb.close()

datos_poniente_Mat =  [
    PROM['2015']['P_M'],
    PROM['2014']['P_M'],
    PROM['2013']['P_M'],
    PROM['2012']['P_M'],
    PROM['2011']['P_M'],
    PROM['2010']['P_M'],
    PROM['2009']['P_M'],
    PROM['2008']['P_M'],
    PROM['2007']['P_M'],
    PROM['2006']['P_M'],
    PROM['2005']['P_M'],
    PROM['2004']['P_M'],
    PROM['2003']['P_M'],
    PROM['2002']['P_M'],
    PROM['2001']['P_M'],
    PROM['2000']['P_M'],
    PROM['1999']['P_M'],
    PROM['1998']['P_M'],
    ]

datos_poniente_Len =  [
    PROM['2015']['P_L'],
    PROM['2014']['P_L'],
    PROM['2013']['P_L'],
    PROM['2012']['P_L'],
    PROM['2011']['P_L'],
    PROM['2010']['P_L'],
    PROM['2009']['P_L'],
    PROM['2008']['P_L'],
    PROM['2007']['P_L'],
    PROM['2006']['P_L'],
    PROM['2005']['P_L'],
    PROM['2004']['P_L'],
    PROM['2003']['P_L'],
    PROM['2002']['P_L'],
    PROM['2001']['P_L'],
    PROM['2000']['P_L'],
    PROM['1999']['P_L'],
    PROM['1998']['P_L'],
    ]

datos_oriente_Mat =  [
    PROM['2015']['O_M'],
    PROM['2014']['O_M'],
    PROM['2013']['O_M'],
    PROM['2012']['O_M'],
    PROM['2011']['O_M'],
    PROM['2010']['O_M'],
    PROM['2009']['O_M'],
    PROM['2008']['O_M'],
    PROM['2007']['O_M'],
    PROM['2006']['O_M'],
    PROM['2005']['O_M'],
    PROM['2004']['O_M'],
    PROM['2003']['O_M'],
    PROM['2002']['O_M'],
    PROM['2001']['O_M'],
    PROM['2000']['O_M'],
    PROM['1999']['O_M'],
    PROM['1998']['O_M'],
    ]

datos_oriente_Len =  [
    PROM['2015']['O_L'],
    PROM['2014']['O_L'],
    PROM['2013']['O_L'],
    PROM['2012']['O_L'],
    PROM['2011']['O_L'],
    PROM['2010']['O_L'],
    PROM['2009']['O_L'],
    PROM['2008']['O_L'],
    PROM['2007']['O_L'],
    PROM['2006']['O_L'],
    PROM['2005']['O_L'],
    PROM['2004']['O_L'],
    PROM['2003']['O_L'],
    PROM['2002']['O_L'],
    PROM['2001']['O_L'],
    PROM['2000']['O_L'],
    PROM['1999']['O_L'],
    PROM['1998']['O_L'],
    ]

fig = plt.figure()
ax = fig.add_subplot(111)
ax.bar(ANIOS,datos_oriente_Len,color = "b",width=0.5)
ax.set_xticks(ANIOS, minor=False)
ax.set_ylabel("Puntajes promedio Lenguaje")
ax.set_xlabel("Oriente anual")
plt.show()

"""
datos = [datos_poniente_Mat ,datos_oriente_Mat ]
X = np.arange(4)
plt.bar(X + 0.00, datos[0], color = "b", width = 0.25)
plt.bar(X + 0.25, datos[1], color = "g", width = 0.25)
plt.xticks(X+0.38,ANIOS)


"""

