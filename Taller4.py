import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
#wb = Workbook()
#ws = wb.active()
listPromL = []
listPromM = []
listPromC = []

wb = load_workbook(filename = 'prueba3.xlsx')
#sheet_ranges = wb['Sheet1']
ws = wb['Sheet1']
#print(ws.max_row)
#print(sheet_ranges['J2'].value)
def ad():
	lista = ['J','K','L','M','N','P','Q','R']
	for x in lista:
		#J K L Region
		ww = ws[x+'1'].value
		#print(ww[0:4])
		if (ww[0:4] == "prom"):
			varInicio = x
			print (varInicio)
		#L M N Provincia
		#P Q R Comunaa Publica
	#return x
def lectura(list,y): #Se le pasa una lista y una columna
	for x in range(2,ws.max_row+1):
		asd = str(y)+ str(x)
		#print(asd)
		list.append(ws[asd].value)
#lectura(listPromL,'J')
#print(listPromL)	
ad()	